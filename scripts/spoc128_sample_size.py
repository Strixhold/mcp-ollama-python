#!/usr/bin/env python3
"""
WWCPI SPOC 128 — Single Sample Plan (Normal), Accept on Zero (C=0)

Run this script once to generate 'spoc128_sample_size_calculator.xlsx'.
Open the file in Excel, go to the Calculator sheet, type in your lot
quantity and pick an AQL level — the sample size appears instantly.
"""

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_FILE = "spoc128_sample_size_calculator.xlsx"

AQL_LEVELS = [
    "0.010", "0.015", "0.025", "0.040", "0.065", "0.10",
    "0.15",  "0.25",  "0.40",  "0.65",  "1.0",   "1.5",
    "2.5",   "4.0",   "6.5",   "10.0",
]

# (min_qty_inclusive, display_label)
LOT_RANGES = [
    (2,      "2–8"),
    (9,      "9–15"),
    (16,     "16–25"),
    (26,     "26–50"),
    (51,     "51–90"),
    (91,     "91–150"),
    (151,    "151–280"),
    (281,    "281–500"),
    (501,    "501–1,200"),
    (1201,   "1,201–3,200"),
    (3201,   "3,201–10,000"),
    (10001,  "10,001–35,000"),
    (35001,  "35,001–150,000"),
    (150001, "150,001–500,000"),
    (500001, "500,001 & Over"),
]

# None  → inspect ALL items (lot too small for sampling at that AQL)
# "?"   → value was not legible in the source document; verify SPOC 128
SAMPLE_TABLE = [
    #.010   .015   .025   .040   .065   .10    .15    .25    .40    .65    1.0    1.5    2.5   4.0   6.5  10.0
    [None,  None,  None,  None,  None,  None,  None,  None,  None,  None,  None,  None,    5,    3,    3,    3],  # 2-8
    [None,  None,  None,  None,  None,  None,  None,  None,  None,  None,   13,    8,    7,    3,    3,    3],  # 9-15
    [None,  None,  None,  None,  None,  None,  None,  None,  None,   20,   13,    8,    7,    3,    3,    3],  # 16-25
    [None,  None,  None,  None,  None,  None,  None,  None,   32,   20,   13,    8,    8,    7,    5,    3],  # 26-50
    [None,  None,  None,  None,  None,  None,   80,   50,   32,   21,   13,   11,   11,    8,    5,    4],  # 51-90
    [None,  None,  None,  None,  None,  125,   80,   50,   32,   22,   13,   13,   11,    9,    6,    5],  # 91-150
    [None,  None,  None,  None,  200,  125,   80,   50,   32,   29,   29,   19,   13,   10,    7,    6],  # 151-280
    [None,  None,  None,  315,  200,  128,   80,   50,   48,   47,   29,   21,   16,   11,    9,    7],  # 281-500
    [None,   800,  500,  315,  200,  125,   80,   75,   73,   47,   34,   27,   19,   15,   11,    8],  # 501-1200
    [1250,   800,  500,  315,  200,  125,  120,  116,   73,   53,   42,   35,   23,   18,   13,    9],  # 1201-3200
    [1250,   800,  500,  315,  200,  192,  189,  116,   86,   68,   50,   38,   29,   22,   15,    9],  # 3201-10000
    [1250,   800,  500,  315,  300,  294,  189,  135,  108,   77,   60,   46,   35,   29,   15,    9],  # 10001-35000
    [1250,   800,  500,  490,  476,  "?",  218,  170,  123,   96,   74,   56,   40,   29,   15,    9],  # 35001-150000
    [1250,   800,  750,  715,  476,  "?",  270,  200,  156,  119,   90,   64,   40,   29,   15,    9],  # 150001-500000
    [1250,  1200, 1112,  715,  558,  "?",  "?",  244,  189,  143,  102,   64,   40,   29,   15,    9],  # 500001+
]

# ── Style helpers ─────────────────────────────────────────────────────────────

def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex6):
    return PatternFill(start_color=hex6, end_color=hex6, fill_type="solid")


def _style(cell, bold=False, fg="000000", bg=None, halign="center",
           size=11, italic=False, wrap=False):
    cell.font = Font(bold=bold, color=fg, size=size, italic=italic)
    if bg:
        cell.fill = _fill(bg)
    cell.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=wrap)
    cell.border = _border()


# ── Reference Table sheet ─────────────────────────────────────────────────────

def build_reference_sheet(wb):
    ws = wb.create_sheet("Reference Table")
    ws.sheet_view.showGridLines = False

    # Row 1 — main title
    ws.merge_cells("A1:R1")
    ws["A1"] = "WWCPI SPOC 128  —  Single Sample Plan (Normal), Accept on Zero (C=0)"
    _style(ws["A1"], bold=True, fg="FFFFFF", bg="1F4E79", size=13)
    ws.row_dimensions[1].height = 24

    # Row 2 — AQL subtitle
    ws.merge_cells("A2:R2")
    ws["A2"] = "Acceptable Quality Level (AQL) — Percent Defective"
    _style(ws["A2"], bold=True, fg="FFFFFF", bg="2E75B6")
    ws.row_dimensions[2].height = 18

    # Row 3 — column headers
    ws["A3"] = "Min Qty"
    _style(ws["A3"], bold=True, fg="FFFFFF", bg="1F4E79", size=10)
    ws["B3"] = "Lot Size"
    _style(ws["B3"], bold=True, fg="FFFFFF", bg="1F4E79")

    for i, aql in enumerate(AQL_LEVELS):
        cell = ws.cell(row=3, column=i + 3, value=float(aql))
        _style(cell, bold=True, fg="FFFFFF", bg="2E75B6")

    ws.row_dimensions[3].height = 18

    # Rows 4-18 — data
    for ri, ((min_qty, label), samples) in enumerate(zip(LOT_RANGES, SAMPLE_TABLE)):
        row = ri + 4
        row_bg = "DEEAF1" if ri % 2 == 0 else "FFFFFF"
        ws.row_dimensions[row].height = 18

        ws.cell(row=row, column=1, value=min_qty)
        _style(ws.cell(row=row, column=1), bg=row_bg, size=10)

        ws.cell(row=row, column=2, value=label)
        _style(ws.cell(row=row, column=2), bold=True, bg=row_bg, halign="left")

        for ci, val in enumerate(samples):
            if val is None:
                display, cell_bg = "ALL", "FFF2CC"
            elif val == "?":
                display, cell_bg = "? *", "FCE4D6"
            else:
                display, cell_bg = val, row_bg
            cell = ws.cell(row=row, column=ci + 3, value=display)
            _style(cell, bg=cell_bg)

    # Footer note
    note_row = len(LOT_RANGES) + 5
    ws.merge_cells(f"A{note_row}:R{note_row}")
    ws[f"A{note_row}"] = (
        "ALL = Inspect 100% of lot  |  "
        "? * = Value not legible in source image — verify against the original SPOC 128 document"
    )
    ws[f"A{note_row}"].font = Font(italic=True, size=9, color="595959")
    ws[f"A{note_row}"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[note_row].height = 16

    # Column widths
    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 19
    for col in range(3, 3 + len(AQL_LEVELS)):
        ws.column_dimensions[get_column_letter(col)].width = 7

    ws.freeze_panes = "C4"


# ── Calculator sheet ──────────────────────────────────────────────────────────

def build_calculator_sheet(wb):
    ws = wb.create_sheet("Calculator", 0)
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("B2:G2")
    ws["B2"] = "SPOC 128 — Sample Size Calculator"
    ws["B2"].font = Font(bold=True, size=16, color="1F4E79")
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 30

    ws.merge_cells("B3:G3")
    ws["B3"] = "Single Sample Plan (Normal)  |  Accept on Zero (C=0)  |  WWCPI Proprietary"
    ws["B3"].font = Font(italic=True, size=10, color="595959")
    ws["B3"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 16

    # Helper: label cell (right-aligned, column B)
    def lbl(row, text):
        c = ws.cell(row=row, column=2, value=text)
        c.font = Font(bold=True, size=11, color="1F4E79")
        c.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[row].height = 26

    # Helper: merged yellow input cell (columns C:E)
    def inp(row):
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
        c = ws.cell(row=row, column=3)
        c.fill = _fill("FFF2CC")
        c.border = _border()
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font = Font(bold=True, size=13)
        return c

    # Row 5 — Quantity input
    lbl(5, "Lot Quantity  →")
    inp(5)  # user types here; left blank

    # Row 6 — AQL dropdown
    lbl(6, "AQL Level  →")
    aql_inp = inp(6)
    aql_inp.value = "1.0"  # sensible default

    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(AQL_LEVELS)}"',
        allow_blank=False,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid AQL",
        error="Please select a value from the dropdown list.",
    )
    dv.sqref = "C6:E6"
    ws.add_data_validation(dv)

    # Row 7 — spacer
    ws.row_dimensions[7].height = 10

    # Row 8 — Sample Size result
    lbl(8, "Sample Size  →")
    ws.merge_cells("C8:E8")
    result = ws["C8"]
    result.fill = _fill("E2EFDA")
    result.border = _border()
    result.alignment = Alignment(horizontal="center", vertical="center")
    result.font = Font(bold=True, size=20, color="375623")
    ws.row_dimensions[8].height = 44

    # VLOOKUP formula:
    #   - column A of Reference Table holds the min lot-size (sorted ascending)
    #   - VLOOKUP with TRUE (approximate match) returns the row where min ≤ qty
    #   - MATCH finds which AQL column to use; +2 offsets past columns A and B
    #   - Nested IFs handle "ALL" and "? *" text values gracefully
    ref = "'Reference Table'"
    tbl = f"{ref}!$A$4:$R$18"
    aql_hdr = f"{ref}!$C$3:$R$3"
    col_idx = f"MATCH(VALUE(C6),{aql_hdr},0)+2"
    vlookup = f"VLOOKUP(C5,{tbl},{col_idx},TRUE)"

    result.value = (
        f'=IF(OR(C5="",C6=""),"← Enter quantity & AQL",'
        f'IFERROR('
        f'IF({vlookup}="ALL","Inspect ALL (100%)",'
        f'IF({vlookup}="? *","⚠ Verify source table",{vlookup})),'
        f'"Invalid input"))'
    )

    # Row 9 — note about ALL
    ws.merge_cells("C9:G9")
    ws["C9"] = (
        '"Inspect ALL" = lot is smaller than the minimum sample at that AQL  |  '
        '"⚠ Verify" = source value was illegible'
    )
    ws["C9"].font = Font(italic=True, size=8, color="595959")
    ws["C9"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[9].height = 14

    # Instructions block
    ws.row_dimensions[11].height = 16
    ws.merge_cells("B11:G11")
    ws["B11"] = "How to use"
    ws["B11"].font = Font(bold=True, size=10, color="1F4E79")
    ws["B11"].alignment = Alignment(horizontal="left", vertical="center")

    steps = [
        "1.  Click the yellow 'Lot Quantity' cell and type the number of parts in the lot.",
        "2.  Click the yellow 'AQL Level' cell and choose a level from the dropdown.",
        "3.  The green cell shows the required sample size automatically.",
        "4.  Three cells in the .10 AQL column (rows 35,001 +) are marked '⚠ Verify' —",
        "     they were not legible in the source image. Cross-check against the original SPOC 128.",
    ]
    for i, text in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:G{r}")
        ws[f"B{r}"] = text
        ws[f"B{r}"].font = Font(size=9, color="404040", italic=(i >= 3))
        ws[f"B{r}"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 15

    # Column widths
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 13
    ws.column_dimensions["D"].width = 13
    ws.column_dimensions["E"].width = 13
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 10


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    build_calculator_sheet(wb)
    build_reference_sheet(wb)

    wb.save(OUTPUT_FILE)
    print(f"Saved: {OUTPUT_FILE}")
    print()
    print("  Sheet 1 'Calculator'      — enter qty + AQL, sample size appears")
    print("  Sheet 2 'Reference Table' — full SPOC 128 C=0 plan")
    print()
    print("  NOTE: Three cells in the AQL 0.10 column (lot sizes 35,001+) and")
    print("  two in the AQL 0.15 column (500,001+) were not readable in the")
    print("  source image and are marked '? *'. Please verify them against")
    print("  the original SPOC 128 document before using those combinations.")


if __name__ == "__main__":
    main()
